#!/usr/bin/env python

# export data sheets from xlsx to csv
# encoding=utf8

from openpyxl import load_workbook
import csv
from os import sys
import codecs
reload(sys)
sys.setdefaultencoding('utf8')
from tkinter import *
from tkinter import filedialog as fd
from tkinter import messagebox

def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file,data_only=True)
    for worksheet_name in sheets:
        #print("Export " + worksheet_name + " ...")

        try:
            
            worksheet = workbook[worksheet_name]
            sheet=workbook.active
            # get max row count
            max_row=sheet.max_row
            # get max column count
            max_column=sheet.max_column
            

        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)
            
        your_csv_file = open(''.join([worksheet_name,'.txt']), 'wb')
        wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        
        title=[]
        title.append("ARTICLE_CODE")
        title.append("BARCODE")
        title.append("PROMO_PRICE")
        title.append("PROMO_TYPE")
        title.append("PROMO_STARTDATE")
        title.append("PROMO_ENDDATE")
        title.append("PROMO_STARTTIME")
        title.append("PROMO_ENDTIME")

        wr.writerow(title)
        
        for i in range(9, max_row+1):
            sdate=sheet['C2'].value
            stdate=sdate.strftime('%Y%m%d')

            edate=sheet['E2'].value
            endate=edate.strftime('%Y%m%d')
            
            lrow = []
            lrow.append(sheet['H'+str(i)].value)  #Gold Code
            lrow.append(sheet['F'+str(i)].value)  #Barcode
            lrow.append(sheet['AJ'+str(i)].value) #PSP
            lrow.append("1") #Promo Type                        
            lrow.append(stdate) #Start Date
            lrow.append(endate) #End Date
            lrow.append("000100") #Start Time
            lrow.append("235900") #End Time                   
                 
            if lrow[0] is not None:
            
            #lrow=sheet['H'+str(i)].value+','+sheet['F'+str(i)].value+','+sheet['AJ'+str(i)].value+',1,'+sheet['C2'].value+','+sheet['E2'].value+',000100,235900'

            #if lrow is not None:
                wr.writerow(lrow)
            
        #print(" ... done")
        messagebox.showinfo("Done", "Convert "+worksheet_name+ " to CSV process is finished!")
        your_csv_file.close()


def get_file():
    ftypes=[('Excel Work Book','*.xlsx')]
    file = fd.askopenfile(mode='rt',filetypes=ftypes)
    if file: 
        #print(file.name)

        sheets = []
        workbook = load_workbook(file.name,read_only=True,data_only=True)
        
        all_worksheets = workbook.sheetnames
        for worksheet_name in all_worksheets:
            sheets.append(worksheet_name)        
        
        csv_from_excel(file.name, sheets)       
        

        filename = worksheet_name+"_final.txt"
        f = codecs.open(filename,encoding='utf-8')
        contents = f.read()

        newcontents = contents.replace('"','')      

        nf=open(filename,"w")
        nf.write(newcontents)

        messagebox.showinfo("Done", "Convert "+worksheet_name+ " to txt is finished!")
        
        f.close()

master = Tk()
master.title("Excel to CSV converter")
Label(master, text="Click Browse to export Excel(xlsx) to CSV.").grid(row=0)


Button(master, text='Quit', command=master.quit).grid(row=3, column=0, sticky=W, pady=4)
Button(master, text='Browse', command=get_file).grid(row=3, column=1, sticky=W, pady=4)    

mainloop( )
